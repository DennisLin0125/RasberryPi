import time
import requests
import openpyxl as xl
from bs4 import BeautifulSoup

def create_dashboard(filename, stocks, token=''): # 初始化時執行一次
    wb=xl.Workbook()
    ws=wb.active
    ws.title='股票看板'
    # 建立第一列的欄名
    ws['A1'].value='股票代號'
    ws['B1'].value='股票名稱'
    ws['C1'].value='開盤價'
    ws['D1'].value='最高價'
    ws['E1'].value='最低價'
    ws['F1'].value='成交價'
    ws['G1'].value='昨日收盤價'
    ws['H1'].value='帳跌幅'
    ws['I1'].value='最近更新'
    ws['J1'].value='Line Notify Token'
    # 將欲觀察的股票代號填入 A2, A3, A4, ... 儲存格
    for idx, stock in enumerate(stocks): 
        ws[f'A{idx + 2}'].value=stock
    # 將 Line Notify token 填入 J2 儲存格
    ws['J2'].value=token
    # 將工作簿存檔
    try:
        wb.save(filename)
    except Exception:
        print(f"Permission Error : {filename} 開啟中, 請關閉")
    return wb, ws

def wearn_crawler(sid):
    res=requests.get(f"https://stock.wearn.com/cdata.asp?kind={sid}")
    res.encoding="big5"
    html=BeautifulSoup(res.text, "html.parser")
    table=html.findAll("table")[0] 
    trs=table.findAll("tr")
    tr=trs[2]
    tds=tr.findAll("td")
    data=[float(td.text.replace("\xa0", "").replace(",", "")) for td in tds[1:]]
    name=trs[0].td.text.split(" ")[1][:-1]
    last_close=float(trs[3].findAll("td")[4].text.replace("\xa0", "").replace(",", ""))
    return {
        "name": name,
        "open": data[0],
        "high": data[1],
        "low": data[2],
        "close": data[3],
        "last": last_close
    }

def notify(msg, token):
    url="https://notify-api.line.me/api/notify"
    headers={"Authorization": "Bearer " + token}
    payload={"message": msg}
    r=requests.post(url, headers=headers, params=payload)
    if r.status_code==requests.codes.ok:   
        return '訊息發送成功！'   
    else:   
        return f'訊息發送失敗: {r.status_code}'   

def update_dashboard(ws):
    last_row=ws.max_row  # 取得總列數
    msg=['']             # 儲存推播訊息之串列
    bigStockURL='https://invest.cnyes.com/index/TWS/TSE01'

    web = requests.get(bigStockURL)  # 取得網頁內容
    soup = BeautifulSoup(web.text, "html.parser")  # 轉換內容

    bigPrice = soup.find('div', class_='jsx-2214436525 info-lp')
    updownPrice = soup.find('div', class_='jsx-2214436525 info-change')
    updownPricent = soup.find('div', class_='jsx-2214436525 change-percent')

    # wsArr.append(bigPrice.get_text())
    # wsArr.append(updownPrice.contents[0].get_text())
    # wsArr.append(updownPricent.get_text())

    msg.append(f'加權指數 {bigPrice.get_text()} ({updownPricent.get_text()})')
    
    token=ws['J2'].value
    for i in range(2, last_row + 1):  # 讀取欲觀察股票清單
        stock_id=ws[f'A{i}'].value
        data=wearn_crawler(stock_id)  # 呼叫爬蟲函式擷取聚財網收盤資料
        print(data)
        ws[f'B{i}'].value=data["name"]
        ws[f'C{i}'].value=data["open"]
        ws[f'D{i}'].value=data["high"]
        ws[f'E{i}'].value=data["low"]
        ws[f'F{i}'].value=data["close"]
        ws[f'G{i}'].value=data["last"]
        delta=round((data["close"]/data["last"] - 1) * 100, 2) # 計算漲跌幅
        ws[f'H{i}'].value=delta
        msg.append(f'{data["name"]} {data["close"]} ({delta}%)')            
    ws['I2'].value=time.strftime("%Y-%m-%d %H:%M:%S") 
    print(f"最近更新時間 : {ws['I2'].value}")
    msg.append(ws['I2'].value)  # 訊息以更新時間結尾
    notify('\n'.join(msg), token)  # 呼叫 Line Notify 函式傳送推播訊息